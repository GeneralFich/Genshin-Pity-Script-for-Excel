import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

ERROR_MESSAGE = {
    FileNotFoundError: ("ERROR: Make sure the Excel workbook is"
                        " in the same folder as this .py file and "
                        " spelled correctly in main.py."),
    ValueError: ("ERROR: Excel workbook appears to be corrupt."),
    KeyError: ("ERROR: Make sure the Excel worksheet with summon data exists"
               " and is spelled correctly in main.py."),
}


class PityCalculator():

    def __init__(self, wb_filepath, worksheet, banner_column, star_column):
        self.filepath = wb_filepath
        self.worksheet = worksheet
        self.banner_column = banner_column
        self.star_column = star_column

        self.new_sheet_name = "Counts_Till_Pity"
        self.pity_count = {5: 90, 4: 10}

        # Read workbook and sheet.
        try:
            self.workbook = load_workbook(self.filepath)
            self.df = pd.read_excel(wb_filepath, sheet_name=self.worksheet)
            # Sort such that rows from same banner are adjacent.
            self.df.sort_values(by=[self.banner_column], inplace=True)
            # Reset the row index numbers.
            self.df.index = range(len(self.df))
            self.banners = self.df[banner_column].unique()
        except (FileNotFoundError, ValueError, KeyError) as e:
            self.handle_exceptions(e)
            return

        # Generate pity data if there was no error reading workbook and sheet.
        self.generate_pity_data()
        print("\nSuccess! Check out the new tab in your Excel workbook: "
              f'{self.filepath} / "{self.worksheet}" tab\n')

    def handle_exceptions(self, e):
        print()
        print(ERROR_MESSAGE.get(e.__class__))
        print("> No changes have been made. Fix the error and try again.")
        print("> Exiting pity calculator...\n")

    def get_last_row_num(self, rows):
        last_row = rows.tail(1)
        try:
            row_num = last_row.index[0]
        except IndexError:
            return IndexError
        return row_num

    def get_count_till_pity(self, banner, star_level):
        max_pity = self.pity_count.get(star_level)

        # Get index of the last ever pull on the banner.
        all_banner_pulls = self.df[self.df[self.banner_column] == banner]
        last_banner_row_index = self.get_last_row_num(all_banner_pulls)
        if last_banner_row_index is IndexError:
            return "N/A: No pull made on banner"

        # Get index of the last ever star_level pull on the banner.
        all_x_star_pulls = self.df[
            (self.df[self.banner_column] == banner)
            & (self.df[self.star_column] == star_level)]
        last_x_star_index = self.get_last_row_num(all_x_star_pulls)
        if last_x_star_index is IndexError:
            return max_pity - len(all_banner_pulls)

        # count until pity = total pity count - (last pull on banner -
        # last pull at star_level on banner)
        count_till_pity = (max_pity -
                           (last_banner_row_index - last_x_star_index))
        return count_till_pity

    def generate_pity_data(self):
        pity_dict = {}
        for banner in self.banners:
            pity_dict[banner] = {
                star_level: self.get_count_till_pity(banner, star_level)
                for star_level in self.pity_count.keys()
            }
        pity_df = pd.DataFrame(pity_dict)

        if self.new_sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[self.new_sheet_name])
        sheet = self.workbook.create_sheet(self.new_sheet_name)
        [sheet.append(row)
         for row in dataframe_to_rows(pity_df, index=True, header=True)]
        self.workbook.save(self.filepath)
